#!/usr/bin/env python3
from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "database" / "demo"

EXPORTS = [
    {
        "label": "monthly_update",
        "db_path": DEMO_DIR / "monthly_update_demo.sqlite3",
        "xlsx_path": DEMO_DIR / "demo_monthly_update_editable.xlsx",
        "title": "Demo 月底更新数据库",
    },
    {
        "label": "dashboard",
        "db_path": DEMO_DIR / "dashboard_demo.sqlite3",
        "xlsx_path": DEMO_DIR / "demo_dashboard_editable.xlsx",
        "title": "Demo 看板数据库",
    },
]

SYSTEM_SHEETS = {"_README", "_TABLES"}


def normalize_excel_value(value: Any) -> Any:
    if isinstance(value, bytes):
        return value.hex()
    return value


def table_names(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        """
    ).fetchall()
    return [row[0] for row in rows]


def table_columns(conn: sqlite3.Connection, table: str) -> list[dict[str, Any]]:
    rows = conn.execute(f"PRAGMA table_info({quote_ident(table)})").fetchall()
    return [
        {
            "cid": row[0],
            "name": row[1],
            "type": row[2],
            "notnull": row[3],
            "default": row[4],
            "pk": row[5],
        }
        for row in rows
    ]


def quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def safe_sheet_name(name: str) -> str:
    for char in "[]:*?/\\":  # Excel-forbidden characters
        name = name.replace(char, "_")
    return name[:31]


def set_sheet_style(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in column_cells[:200]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write_readme(wb: Workbook, export: dict[str, Any], table_count: int) -> None:
    ws = wb.active
    ws.title = "_README"
    rows = [
        ["项目", export["title"]],
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["来源数据库", str(export["db_path"])],
        ["表数量", table_count],
        ["怎么改", "可以改表里的普通内容；不要删 sheet、不要改第一行表头、不要删除 id。"],
        ["新增行", "如果要新增记录，先留空 id 或写清楚备注；同步前我会检查。"],
        ["删除行", "不要直接删除整张表；如果想删除某条记录，可以先在旁边备注，或者告诉我。"],
        ["同步规则", "你改完 Excel 后发我，我会先备份 Demo SQLite，再把修改同步回对应 Demo 数据库。"],
        ["注意", "正式数据库和测试数据库不会自动跟着改。"],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_one(export: dict[str, Any]) -> None:
    db_path = Path(export["db_path"])
    xlsx_path = Path(export["xlsx_path"])
    if not db_path.exists():
        raise FileNotFoundError(db_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    tables = table_names(conn)

    wb = Workbook()
    write_readme(wb, export, len(tables))

    tables_ws = wb.create_sheet("_TABLES")
    tables_ws.append(["sheet_name", "table_name", "row_count", "primary_key_columns"])

    for table in tables:
        cols = table_columns(conn, table)
        col_names = [col["name"] for col in cols]
        pk_cols = [col["name"] for col in cols if col["pk"]]
        order_by = ", ".join(quote_ident(col) for col in pk_cols) if pk_cols else "rowid"
        rows = conn.execute(
            f"SELECT * FROM {quote_ident(table)} ORDER BY {order_by}"
        ).fetchall()

        sheet_name = safe_sheet_name(table)
        ws = wb.create_sheet(sheet_name)
        ws.append(col_names)
        for row in rows:
            ws.append([normalize_excel_value(row[col]) for col in col_names])
        set_sheet_style(ws)
        tables_ws.append([sheet_name, table, len(rows), ", ".join(pk_cols)])

    set_sheet_style(tables_ws)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    conn.close()

    # Verify the workbook can be reopened.
    loaded = load_workbook(xlsx_path, read_only=True)
    missing = SYSTEM_SHEETS - set(loaded.sheetnames)
    if missing:
        raise RuntimeError(f"{xlsx_path} 缺少系统 sheet: {sorted(missing)}")
    loaded.close()


def main() -> None:
    for export in EXPORTS:
        export_one(export)
        print(export["xlsx_path"])


if __name__ == "__main__":
    main()
