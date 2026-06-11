#!/usr/bin/env python3
import argparse
import shutil
import sqlite3
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DB_PATH = Path.home() / "Library/Application Support/com.emilylu.financial-planning/financial_planning.sqlite3"
EXPORT_DIR = Path("/private/tmp/financial_planning_numbers_export")
XLSX_2025 = EXPORT_DIR / "2025_收入支出储蓄记录.xlsx"
XLSX_2026 = EXPORT_DIR / "2026_收入支出储蓄记录.xlsx"

HISTORY_START = "2025-01"
HISTORY_END = "2026-04"
ASSET_START = "2025-05"
ASSET_END = "2026-04"

SKIP_EXPENSE_ROWS = {
    "年度支出记录📝",
    "支出总和",
    "刚性支出",
    "弹性支出",
    "外出吃饭次数",
    "和朋友吃次数（非单项）",
}
SKIP_INCOME_ROWS = {"收入明细记录", "总计收入", "总支出", "储蓄率", "合计"}
SKIP_ASSET_ROW_KEYWORDS = ("总额", "信用卡")
ASSET_TOTAL_ROW = "资产总额"
CREDIT_CARD_BILLED_ROW = "信用卡本月已出账单"
CREDIT_CARD_PREVIOUS_ROW = "信用卡上个月没出账单"
CREDIT_CARD_UNBILLED_ROW = "信用卡本月未出账单"

CATEGORY_ALIASES = {
    ("income", "红包🧧"): "红包",
    ("income", "其它"): "其它收入",
    ("income", "其他"): "其它收入",
}

ASSET_NAME_ALIASES = {
    "纳斯达克": "华宝纳斯达克a",
    "华宝纳斯达克": "华宝纳斯达克a",
    "天虹标普": "天虹标普a",
    "摩根标普": "摩根标普a",
    "红利低波": "南方红利低波",
}


def sha256(value: str) -> str:
    import hashlib

    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def make_id(prefix: str, seed: str) -> str:
    return f"{prefix}_{sha256(seed)[:24]}"


def month_end(period: str) -> str:
    year, month = map(int, period.split("-"))
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    last_day = (next_month - datetime.resolution).day
    return f"{period}-{last_day:02d}"


def normalize_amount(value) -> float:
    if value is None or value == "":
        return 0.0
    return abs(float(str(value).replace(",", "")))


def ensure_category(conn: sqlite3.Connection, name: str, kind: str) -> str:
    canonical = CATEGORY_ALIASES.get((kind, name), name).strip()
    row = conn.execute(
        "select id from categories where name = ? and category_kind = ?",
        (canonical, kind),
    ).fetchone()
    if row:
        return row[0]
    category_id = make_id("cat", f"numbers|{kind}|{canonical}")
    rigidity = "flexible" if kind == "expense" else None
    conn.execute(
        """
        insert into categories (
          id, name, category_kind, rigidity, is_personal, is_active,
          sort_order, is_auto_created, source, created_from_raw_category, note
        )
        values (?, ?, ?, ?, 1, 1, 999, 1, 'numbers_history', ?, 'Numbers 历史迁移自动新增，请后续复核')
        on conflict(name, category_kind) do nothing
        """,
        (category_id, canonical, kind, rigidity, name),
    )
    return conn.execute(
        "select id from categories where name = ? and category_kind = ?",
        (canonical, kind),
    ).fetchone()[0]


def asset_guess(name: str) -> tuple[str, str, str]:
    if "债券" in name:
        return "bond_fund", "asset_cat_bond", "asset_sub_bond_fund"
    if "黄金" in name:
        return "gold", "asset_cat_gold", "asset_sub_gold"
    if "红利" in name:
        return "fund", "asset_cat_dividend_low_vol", "asset_sub_dividend_low_vol"
    if "现金" in name:
        return "cash_account", "asset_cat_cash", "asset_sub_cash"
    if "纳斯达克" in name or "信息科技" in name or "科技" in name:
        return "fund", "asset_cat_us_equity", "asset_sub_nasdaq"
    if "标普" in name:
        return "fund", "asset_cat_us_equity", "asset_sub_sp500"
    return "fund", "asset_cat_us_equity", "asset_sub_us_market"


def canonical_asset_name(name: str) -> str:
    return ASSET_NAME_ALIASES.get(name.strip(), name.strip())


def is_real_asset_row(name: str) -> bool:
    clean = name.strip()
    if not clean or clean == "资产-金额":
        return False
    return not any(keyword in clean for keyword in SKIP_ASSET_ROW_KEYWORDS)


def find_exact_subset(rows: list[tuple[str, float]], target: float) -> list[str] | None:
    target_cents = round(target * 100)
    options = [(name, round(amount * 100)) for name, amount in rows if amount > 0]
    matches: list[list[str]] = []

    def walk(index: int, remaining: int, picked: list[str]):
        if len(matches) > 1:
            return
        if remaining == 0:
            matches.append(picked.copy())
            return
        if remaining < 0 or index >= len(options):
            return
        name, cents = options[index]
        walk(index + 1, remaining - cents, picked + [name])
        walk(index + 1, remaining, picked)

    walk(0, target_cents, [])
    return matches[0] if len(matches) == 1 else None


def asset_period_exclusions(workbook_path: Path, start: str, end: str) -> dict[str, set[str]]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["资产收益率_占比计算表 - 资产-金额"]
    rows = row_lookup(ws)
    headers = period_headers_by_dates(ws, start, end)
    exclusions: dict[str, set[str]] = {}
    for col, period in headers.items():
        details: list[tuple[str, float]] = []
        for name, row_idx in rows.items():
            if not is_real_asset_row(name):
                continue
            value = normalize_amount(ws.cell(row_idx, col).value)
            if value:
                details.append((canonical_asset_name(name), value))

        detail_total = round(sum(value for _, value in details), 2)
        official_total = normalize_amount(ws.cell(rows[ASSET_TOTAL_ROW], col).value)
        billed = normalize_amount(ws.cell(rows.get(CREDIT_CARD_BILLED_ROW, 0), col).value) if rows.get(CREDIT_CARD_BILLED_ROW) else 0.0
        previous = normalize_amount(ws.cell(rows.get(CREDIT_CARD_PREVIOUS_ROW, 0), col).value) if rows.get(CREDIT_CARD_PREVIOUS_ROW) else 0.0
        unbilled = normalize_amount(ws.cell(rows.get(CREDIT_CARD_UNBILLED_ROW, 0), col).value) if rows.get(CREDIT_CARD_UNBILLED_ROW) else 0.0
        credit_net = round(-billed - unbilled + previous, 2)
        target_gross = round(official_total - credit_net, 2)
        extra = round(detail_total - target_gross, 2)
        if abs(extra) <= 0.01:
            continue
        if extra < -0.01:
            raise RuntimeError(f"{period} Numbers 资产明细小于资产总额倒推值：明细 {detail_total:.2f}，目标 {target_gross:.2f}")
        excluded = find_exact_subset(details, extra)
        if not excluded:
            raise RuntimeError(f"{period} 无法自动定位资产总额差异 {extra:.2f} 对应的资产行")
        exclusions[period] = set(excluded)
    return exclusions


def ensure_asset(conn: sqlite3.Connection, name: str) -> str:
    canonical = canonical_asset_name(name)
    row = conn.execute("select id from assets where name = ?", (canonical,)).fetchone()
    if row:
        return row[0]
    asset_type, main_id, sub_id = asset_guess(canonical)
    asset_id = make_id("asset", f"numbers|{canonical}")
    conn.execute(
        """
        insert into assets (
          id, name, asset_type, main_asset_category_id, sub_asset_category_id,
          currency, platform, is_dca, status, note
        )
        values (?, ?, ?, ?, ?, 'CNY', 'Numbers 历史迁移', 0, 'active', '从 Numbers 历史数据迁移创建')
        on conflict(name, platform) do nothing
        """,
        (asset_id, canonical, asset_type, main_id, sub_id),
    )
    return conn.execute("select id from assets where name = ?", (canonical,)).fetchone()[0]


def period_headers_by_dates(ws, start: str, end: str, header_row: int = 2) -> dict[int, str]:
    headers = {}
    for col in range(2, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if isinstance(value, datetime):
            period = value.strftime("%Y-%m")
            if start <= period <= end:
                headers[col] = period
    return headers


def fixed_year_headers(ws, year: int, start_month: int, end_month: int) -> dict[int, str]:
    headers = {}
    for col, month in enumerate(range(start_month, end_month + 1), start=2):
        if col <= ws.max_column:
            headers[col] = f"{year}-{month:02d}"
    return headers


def row_lookup(ws) -> dict[str, int]:
    return {str(ws.cell(row, 1).value).strip(): row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value}


def authoritative_totals() -> dict[tuple[str, str], float]:
    totals: dict[tuple[str, str], float] = {}
    wb25 = load_workbook(XLSX_2025, data_only=True)
    ws25_exp = wb25["支出明细表 - 年度支出记录📝"]
    rows25_exp = row_lookup(ws25_exp)
    for col, period in period_headers_by_dates(ws25_exp, "2025-01", "2025-12").items():
        totals[("expense", period)] = normalize_amount(ws25_exp.cell(rows25_exp["支出总和"], col).value)
    ws25_inc = wb25["收入明细表 - 收入明细记录"]
    rows25_inc = row_lookup(ws25_inc)
    for col, period in period_headers_by_dates(ws25_inc, "2025-06", "2025-12").items():
        totals[("income", period)] = normalize_amount(ws25_inc.cell(rows25_inc["总计收入"], col).value)

    wb26 = load_workbook(XLSX_2026, data_only=True)
    ws26_exp = wb26["支出明细表 - 年度支出记录📝"]
    rows26_exp = row_lookup(ws26_exp)
    for col, period in fixed_year_headers(ws26_exp, 2026, 1, 4).items():
        totals[("expense", period)] = normalize_amount(ws26_exp.cell(rows26_exp["支出总和"], col).value)
    ws26_inc = wb26["收入明细表 - 收入明细记录"]
    rows26_inc = row_lookup(ws26_inc)
    for col, period in period_headers_by_dates(ws26_inc, "2026-01", "2026-04").items():
        totals[("income", period)] = normalize_amount(ws26_inc.cell(rows26_inc["总计收入"], col).value)
    return totals


def import_month_summary(
    conn: sqlite3.Connection,
    workbook_path: Path,
    sheet_name: str,
    kind: str,
    headers: dict[int, str],
    skip_names: set[str] | None,
    stats: Counter,
):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    for row_idx in range(3, ws.max_row + 1):
        raw_name = ws.cell(row_idx, 1).value
        if not raw_name:
            continue
        name = str(raw_name).strip()
        if skip_names and name in skip_names:
            continue
        for col, period in headers.items():
            amount = normalize_amount(ws.cell(row_idx, col).value)
            if amount == 0:
                continue
            category_id = ensure_category(conn, name, kind)
            confirmed_id = make_id("confirmed", f"numbers_history|{kind}|{period}|{name}")
            conn.execute(
                """
                insert into confirmed_transactions (
                  id, source_kind, raw_transaction_id, period_month, transaction_date,
                  transaction_type, amount, currency, category_id, raw_category_snapshot,
                  include_in_stats, confirmation_status, adjustment_reason, note
                )
                values (?, 'historical_numbers_summary', null, ?, ?, ?, ?, 'CNY', ?, ?, 1, 'confirmed', 'Numbers 历史月汇总迁移', '按 Numbers 月汇总导入')
                on conflict(id) do update set
                  amount = excluded.amount,
                  category_id = excluded.category_id,
                  raw_category_snapshot = excluded.raw_category_snapshot,
                  updated_at = current_timestamp
                """,
                (confirmed_id, period, month_end(period), kind, amount, category_id, name),
            )
            stats[f"{kind}_summary_rows"] += 1


def apply_authoritative_adjustments(conn: sqlite3.Connection, stats: Counter):
    totals = authoritative_totals()
    adjustment_category_ids = {
        "expense": ensure_category(conn, "Numbers校准调整", "expense"),
        "income": ensure_category(conn, "Numbers校准调整", "income"),
    }
    for (kind, period), expected_amount in sorted(totals.items()):
        current_amount = conn.execute(
            """
            select coalesce(round(sum(amount), 2), 0)
            from confirmed_transactions
            where period_month = ? and transaction_type = ?
              and include_in_stats = 1 and confirmation_status = 'confirmed'
            """,
            (period, kind),
        ).fetchone()[0]
        diff = round(expected_amount - float(current_amount or 0), 2)
        if abs(diff) <= 0.01:
            continue
        adjustment_id = make_id("confirmed", f"numbers_adjustment|{kind}|{period}")
        conn.execute(
            """
            insert into confirmed_transactions (
              id, source_kind, raw_transaction_id, period_month, transaction_date,
              transaction_type, amount, currency, category_id, raw_category_snapshot,
              include_in_stats, confirmation_status, adjustment_reason, note
            )
            values (?, 'historical_numbers_adjustment', null, ?, ?, ?, ?, 'CNY', ?, 'Numbers校准调整', 1, 'confirmed', 'Numbers 总计行校准', ?)
            on conflict(id) do update set
              amount = excluded.amount,
              category_id = excluded.category_id,
              note = excluded.note,
              updated_at = current_timestamp
            """,
            (
                adjustment_id,
                period,
                month_end(period),
                kind,
                diff,
                adjustment_category_ids[kind],
                f"分类合计 {float(current_amount or 0):.2f}；Numbers 总计 {expected_amount:.2f}；差额 {diff:.2f}",
            ),
        )
        stats[f"{kind}_adjustment_rows"] += 1


def import_income_expense(conn: sqlite3.Connection, stats: Counter):
    wb25 = load_workbook(XLSX_2025, data_only=True)
    wb26 = load_workbook(XLSX_2026, data_only=True)

    import_month_summary(
        conn,
        XLSX_2025,
        "支出明细表 - 年度支出记录📝",
        "expense",
        period_headers_by_dates(wb25["支出明细表 - 年度支出记录📝"], "2025-01", "2025-12"),
        SKIP_EXPENSE_ROWS,
        stats,
    )
    # 2026 expense sheet exported headers can retain 2025 labels, so use fixed columns B-E.
    import_month_summary(
        conn,
        XLSX_2026,
        "支出明细表 - 年度支出记录📝",
        "expense",
        fixed_year_headers(wb26["支出明细表 - 年度支出记录📝"], 2026, 1, 4),
        SKIP_EXPENSE_ROWS,
        stats,
    )
    import_month_summary(
        conn,
        XLSX_2025,
        "收入明细表 - 收入明细记录",
        "income",
        period_headers_by_dates(wb25["收入明细表 - 收入明细记录"], "2025-01", "2025-12"),
        SKIP_INCOME_ROWS,
        stats,
    )
    import_month_summary(
        conn,
        XLSX_2026,
        "收入明细表 - 收入明细记录",
        "income",
        period_headers_by_dates(wb26["收入明细表 - 收入明细记录"], "2026-01", "2026-04"),
        SKIP_INCOME_ROWS,
        stats,
    )
    apply_authoritative_adjustments(conn, stats)


def import_asset_snapshots(conn: sqlite3.Connection, stats: Counter):
    for workbook_path, start, end in [(XLSX_2025, "2025-05", "2025-12"), (XLSX_2026, "2026-01", "2026-04")]:
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb["资产收益率_占比计算表 - 资产-金额"]
        headers = period_headers_by_dates(ws, start, end)
        exclusions = asset_period_exclusions(workbook_path, start, end)
        for row_idx in range(3, ws.max_row + 1):
            raw_name = ws.cell(row_idx, 1).value
            if not raw_name or not is_real_asset_row(str(raw_name)):
                continue
            asset_name = canonical_asset_name(str(raw_name))
            for col, period in headers.items():
                if asset_name in exclusions.get(period, set()):
                    stats["asset_snapshot_excluded_by_numbers_total"] += 1
                    continue
                amount = normalize_amount(ws.cell(row_idx, col).value)
                if amount == 0:
                    continue
                asset_id = ensure_asset(conn, asset_name)
                snapshot_id = make_id("asset_snapshot", f"{asset_id}|{period}|v1")
                conn.execute(
                    """
                    insert into monthly_asset_snapshots (
                      id, asset_id, period_month, snapshot_date, original_amount,
                      currency, fx_rate_to_cny, amount_cny, status, version_no, note
                    )
                    values (?, ?, ?, ?, ?, 'CNY', 1, ?, 'held', 1, 'Numbers 历史迁移')
                    on conflict(asset_id, period_month, version_no) do update set
                      snapshot_date = excluded.snapshot_date,
                      original_amount = excluded.original_amount,
                      amount_cny = excluded.amount_cny,
                      status = excluded.status,
                      note = excluded.note,
                      updated_at = current_timestamp
                    """,
                    (snapshot_id, asset_id, period, month_end(period), amount, amount),
                )
                stats["asset_snapshots"] += 1


def cashflow_sheets():
    for path in [XLSX_2025, XLSX_2026]:
        wb = load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            if "每日现金流" in sheet and "Drawings" not in sheet and "总" not in sheet:
                yield path, sheet


def import_investment_cashflows(conn: sqlite3.Connection, stats: Counter):
    exclusions_by_workbook = {
        XLSX_2025: asset_period_exclusions(XLSX_2025, "2025-05", "2025-12"),
        XLSX_2026: asset_period_exclusions(XLSX_2026, "2026-01", "2026-04"),
    }
    for path, sheet in cashflow_sheets():
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet]
        workbook_exclusions = exclusions_by_workbook.get(path, {})
        asset_names = {
            col: canonical_asset_name(str(ws.cell(2, col).value).strip())
            for col in range(2, ws.max_column + 1)
            if ws.cell(2, col).value and is_real_asset_row(str(ws.cell(2, col).value))
        }
        first_date_row_seen = False
        for row_idx in range(3, ws.max_row + 1):
            first = ws.cell(row_idx, 1).value
            if not isinstance(first, datetime):
                continue
            period = first.strftime("%Y-%m")
            if not (ASSET_START <= period <= ASSET_END):
                continue
            if not first_date_row_seen:
                first_date_row_seen = True
                continue
            next_first = ws.cell(row_idx + 1, 1).value if row_idx + 1 <= ws.max_row else None
            if isinstance(next_first, str) and "年化" in next_first:
                continue
            for col, asset_name in asset_names.items():
                value = ws.cell(row_idx, col).value
                if value is None:
                    continue
                signed_amount = float(value)
                if abs(signed_amount) < 0.000001:
                    continue
                if asset_name in workbook_exclusions.get(period, set()):
                    stats["investment_cashflows_excluded_by_numbers_total"] += 1
                    continue
                flow_type = "buy" if signed_amount < 0 else "sell"
                amount = abs(signed_amount)
                asset_id = ensure_asset(conn, asset_name)
                flow_date = first.strftime("%Y-%m-%d")
                flow_id = make_id("cashflow", f"numbers_history|{asset_id}|{flow_date}|{signed_amount:.4f}|{flow_type}")
                conn.execute(
                    """
                    insert into investment_cashflows (
                      id, asset_id, period_month, flow_date, flow_type, amount,
                      currency, fx_rate_to_cny, amount_cny, source_kind, dca_plan_id, note
                    )
                    values (?, ?, ?, ?, ?, ?, 'CNY', 1, ?, 'historical_numbers_cashflow', null, 'Numbers 每日现金流迁移')
                    on conflict(id) do update set
                      amount = excluded.amount,
                      amount_cny = excluded.amount_cny,
                      flow_type = excluded.flow_type,
                      updated_at = current_timestamp
                    """,
                    (flow_id, asset_id, period, flow_date, flow_type, amount, amount),
                )
                stats["investment_cashflows"] += 1


def ensure_credit_card(conn: sqlite3.Connection) -> str:
    row = conn.execute("select id from credit_cards where name = '信用卡' order by created_at limit 1").fetchone()
    if row:
        return row[0]
    card_id = make_id("credit_card", "信用卡")
    conn.execute(
        """
        insert into credit_cards (id, name, institution, note, is_active)
        values (?, '信用卡', null, 'Numbers 历史迁移创建', 1)
        """,
        (card_id,),
    )
    return card_id


def import_credit_cards(conn: sqlite3.Connection, stats: Counter):
    card_id = ensure_credit_card(conn)
    for workbook_path, start, end in [(XLSX_2025, "2025-05", "2025-12"), (XLSX_2026, "2026-01", "2026-04")]:
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb["资产收益率_占比计算表 - 资产-金额"]
        headers = period_headers_by_dates(ws, start, end)
        row_by_name = row_lookup(ws)
        billed_row = row_by_name.get("信用卡本月已出账单")
        previous_row = row_by_name.get("信用卡上个月没出账单")
        unbilled_row = row_by_name.get("信用卡本月未出账单")
        if not billed_row or not previous_row or not unbilled_row:
            continue
        for col, period in headers.items():
            billed = normalize_amount(ws.cell(billed_row, col).value)
            previous = normalize_amount(ws.cell(previous_row, col).value)
            unbilled = normalize_amount(ws.cell(unbilled_row, col).value)
            if billed == 0 and previous == 0 and unbilled == 0:
                continue
            net_adjustment = round(-billed - unbilled + previous, 2)
            entry_id = make_id("credit_month", f"{card_id}|{period}")
            conn.execute(
                """
                insert into monthly_credit_card_entries (
                  id, credit_card_id, period_month, billed_amount, unbilled_amount,
                  previous_unbilled_amount, previous_unbilled_override,
                  previous_unbilled_override_reason, net_adjustment, note, confirmed
                )
                values (?, ?, ?, ?, ?, ?, 0, null, ?, 'Numbers 历史迁移', 1)
                on conflict(credit_card_id, period_month) do update set
                  billed_amount = excluded.billed_amount,
                  unbilled_amount = excluded.unbilled_amount,
                  previous_unbilled_amount = excluded.previous_unbilled_amount,
                  net_adjustment = excluded.net_adjustment,
                  note = excluded.note,
                  confirmed = 1,
                  updated_at = current_timestamp
                """,
                (entry_id, card_id, period, billed, unbilled, previous, net_adjustment),
            )
            stats["credit_card_months"] += 1


def clear_history(conn: sqlite3.Connection):
    conn.execute("delete from confirmed_transactions where period_month between ? and ?", (HISTORY_START, HISTORY_END))
    conn.execute("delete from raw_transactions where transaction_date between '2025-01-01' and '2026-04-30'")
    conn.execute("delete from import_batches where source_type = 'shark_csv_history'")
    conn.execute("delete from monthly_asset_snapshots where period_month between ? and ?", (ASSET_START, ASSET_END))
    conn.execute("delete from investment_cashflows where period_month between ? and ?", (ASSET_START, ASSET_END))
    conn.execute("delete from monthly_credit_card_entries where period_month between ? and ?", (HISTORY_START, HISTORY_END))
    conn.execute("delete from monthly_step_status where period_month between ? and ?", (HISTORY_START, HISTORY_END))
    conn.execute("delete from monthly_closes where period_month between ? and ?", (HISTORY_START, HISTORY_END))


def mark_history_months(conn: sqlite3.Connection, stats: Counter):
    months = [f"2025-{month:02d}" for month in range(1, 13)] + [f"2026-{month:02d}" for month in range(1, 5)]
    for period in months:
        for step in ["import", "expense", "income", "assets", "creditCard", "final"]:
            conn.execute(
                """
                insert into monthly_step_status (period_month, step_key, completed, completed_at)
                values (?, ?, 1, current_timestamp)
                on conflict(period_month, step_key) do update set
                  completed = 1,
                  completed_at = current_timestamp,
                  updated_at = current_timestamp
                """,
                (period, step),
            )
        close_id = make_id("monthly_close", period)
        conn.execute(
            """
            insert into monthly_closes (id, period_month, close_date, status, version_no, confirmed_at, note)
            values (?, ?, ?, 'historical_numbers_imported', 1, current_timestamp, 'Numbers 历史数据迁移完成')
            on conflict(period_month) do update set
              status = excluded.status,
              confirmed_at = current_timestamp,
              note = excluded.note,
              updated_at = current_timestamp
            """,
            (close_id, period, month_end(period)),
        )
        stats["history_months"] += 1


def run(apply: bool, db_path: Path) -> Counter:
    conn = sqlite3.connect(db_path)
    conn.execute("pragma foreign_keys = on")
    stats = Counter()
    try:
        if apply:
            clear_history(conn)
        import_income_expense(conn, stats)
        import_asset_snapshots(conn, stats)
        import_investment_cashflows(conn, stats)
        import_credit_cards(conn, stats)
        mark_history_months(conn, stats)
        if apply:
            conn.commit()
        else:
            conn.rollback()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return stats


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--db", default=str(DB_PATH))
    args = parser.parse_args()
    db_path = Path(args.db)
    if not args.apply:
        preview_path = Path("/private/tmp/financial_planning_numbers_history_preview.sqlite3")
        shutil.copy2(db_path, preview_path)
        db_path = preview_path
    if args.apply:
        backup = db_path.with_name(f"{db_path.stem}_backup_before_numbers_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}{db_path.suffix}")
        shutil.copy2(db_path, backup)
        print(f"backup={backup}")
    stats = run(args.apply, db_path)
    for key in sorted(stats):
        print(f"{key}={stats[key]}")


if __name__ == "__main__":
    main()
