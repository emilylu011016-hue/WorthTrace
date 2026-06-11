#!/usr/bin/env python3
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DB = Path.home() / "Library/Application Support/com.emilylu.financial-planning/financial_planning.sqlite3"
EXPORT_DIR = Path("/private/tmp/financial_planning_numbers_export")
XLSX_2025 = EXPORT_DIR / "2025_收入支出储蓄记录.xlsx"
XLSX_2026 = EXPORT_DIR / "2026_收入支出储蓄记录.xlsx"

TOLERANCE = 0.02


def amount(value) -> float:
    if value is None or value == "":
        return 0.0
    return round(abs(float(value)), 2)


def fixed_header(ws, year: int, start: int, end: int):
    return {f"{year}-{month:02d}": col for col, month in enumerate(range(start, end + 1), start=2) if col <= ws.max_column}


def date_header(ws, start: str, end: str):
    out = {}
    for col in range(2, ws.max_column + 1):
        value = ws.cell(2, col).value
        if isinstance(value, datetime):
            period = value.strftime("%Y-%m")
            if start <= period <= end:
                out[period] = col
    return out


def row_map(ws):
    return {str(ws.cell(row, 1).value).strip(): row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value}


def numbers_expected():
    expected = {"income": {}, "expense": {}, "asset_gross": {}, "credit_net": {}, "asset_total": {}}
    wb25 = load_workbook(XLSX_2025, data_only=True)
    ws25_exp = wb25["支出明细表 - 年度支出记录📝"]
    rows25_exp = row_map(ws25_exp)
    for period, col in date_header(ws25_exp, "2025-01", "2025-12").items():
        expected["expense"][period] = amount(ws25_exp.cell(rows25_exp["支出总和"], col).value)
    ws25_inc = wb25["收入明细表 - 收入明细记录"]
    rows25_inc = row_map(ws25_inc)
    for period, col in date_header(ws25_inc, "2025-06", "2025-12").items():
        expected["income"][period] = amount(ws25_inc.cell(rows25_inc["总计收入"], col).value)

    wb26 = load_workbook(XLSX_2026, data_only=True)
    ws26_exp = wb26["支出明细表 - 年度支出记录📝"]
    rows26_exp = row_map(ws26_exp)
    for period, col in fixed_header(ws26_exp, 2026, 1, 4).items():
        expected["expense"][period] = amount(ws26_exp.cell(rows26_exp["支出总和"], col).value)
    ws26_inc = wb26["收入明细表 - 收入明细记录"]
    rows26_inc = row_map(ws26_inc)
    for period, col in date_header(ws26_inc, "2026-01", "2026-04").items():
        expected["income"][period] = amount(ws26_inc.cell(rows26_inc["总计收入"], col).value)

    for wb, start, end in [(wb25, "2025-05", "2025-12"), (wb26, "2026-01", "2026-04")]:
        ws = wb["资产收益率_占比计算表 - 资产-金额"]
        rows = row_map(ws)
        for period, col in date_header(ws, start, end).items():
            asset_total = amount(ws.cell(rows["资产总额"], col).value)
            billed = amount(ws.cell(rows.get("信用卡本月已出账单", 0), col).value) if "信用卡本月已出账单" in rows else 0.0
            unbilled = amount(ws.cell(rows.get("信用卡本月未出账单", 0), col).value) if "信用卡本月未出账单" in rows else 0.0
            previous = amount(ws.cell(rows.get("信用卡上个月没出账单", 0), col).value) if "信用卡上个月没出账单" in rows else 0.0
            expected["asset_total"][period] = asset_total
            expected["asset_gross"][period] = round(asset_total - (-billed - unbilled + previous), 2)
            if billed or unbilled or previous:
                expected["credit_net"][period] = round(-billed - unbilled + previous, 2)
    return expected


def db_actual():
    conn = sqlite3.connect(DB)
    actual = {"income": {}, "expense": {}, "asset_gross": {}, "credit_net": {}, "asset_total": {}}
    for period, kind, total in conn.execute(
        """
        select period_month, transaction_type, round(sum(amount), 2)
        from confirmed_transactions
        where period_month between '2025-01' and '2026-04'
          and include_in_stats = 1 and confirmation_status = 'confirmed'
        group by period_month, transaction_type
        """
    ):
        actual[kind][period] = float(total or 0)
    for period, total in conn.execute(
        """
        select period_month, round(sum(amount_cny), 2)
        from monthly_asset_snapshots
        where period_month between '2025-05' and '2026-04'
          and version_no = 1 and status = 'held'
        group by period_month
        """
    ):
        actual["asset_gross"][period] = float(total or 0)
    for period, total in conn.execute(
        """
        select period_month, round(sum(net_adjustment), 2)
        from monthly_credit_card_entries
        where period_month between '2025-01' and '2026-04'
        group by period_month
        """
    ):
        actual["credit_net"][period] = float(total or 0)
    for period in sorted(set(actual["asset_gross"]) | set(actual["credit_net"])):
        actual["asset_total"][period] = round(actual["asset_gross"].get(period, 0.0) + actual["credit_net"].get(period, 0.0), 2)
    conn.close()
    return actual


def main():
    expected = numbers_expected()
    actual = db_actual()
    mismatches = []
    for group in ["income", "expense", "asset_gross", "credit_net", "asset_total"]:
        periods = sorted(set(expected[group]) | set(actual[group]))
        for period in periods:
            exp = round(expected[group].get(period, 0.0), 2)
            act = round(actual[group].get(period, 0.0), 2)
            if abs(exp - act) > TOLERANCE:
                mismatches.append((group, period, exp, act, round(act - exp, 2)))
    if mismatches:
        print("MISMATCH")
        for row in mismatches:
            print("|".join(map(str, row)))
        raise SystemExit(1)
    print("OK")
    for group in ["income", "expense", "asset_gross", "credit_net", "asset_total"]:
        print(group, len(expected[group]))


if __name__ == "__main__":
    main()
