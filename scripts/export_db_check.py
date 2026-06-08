#!/usr/bin/env python3
import sqlite3
import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from validate_numbers_db import TOLERANCE, db_actual, numbers_expected


DB = Path.home() / "Library/Application Support/com.emilylu.financial-planning/financial_planning.sqlite3"
DEFAULT_OUT = Path(__file__).resolve().parents[1] / "financial_planning_db_check.xlsx"

QUERIES = [
    (
        "01_数据源总览",
        """
        select 'confirmed_transactions' as 表名, source_kind as 来源, min(period_month) as 最早月份, max(period_month) as 最晚月份, count(*) as 条数
        from confirmed_transactions group by source_kind
        union all
        select 'investment_cashflows', source_kind, min(period_month), max(period_month), count(*)
        from investment_cashflows group by source_kind
        union all
        select 'monthly_closes', status, min(period_month), max(period_month), count(*)
        from monthly_closes group by status
        order by 表名, 来源
        """,
    ),
    (
        "02_月度收支汇总",
        """
        select
          period_month as 月份,
          round(sum(case when transaction_type='income' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end), 2) as 收入,
          round(sum(case when transaction_type='expense' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end), 2) as 支出,
          round(sum(case when transaction_type='income' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end) -
                sum(case when transaction_type='expense' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end), 2) as 储蓄金额,
          case when sum(case when transaction_type='income' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end) > 0
            then round((sum(case when transaction_type='income' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end) -
                sum(case when transaction_type='expense' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end)) /
                sum(case when transaction_type='income' and include_in_stats=1 and confirmation_status='confirmed' then amount else 0 end), 4)
            else null end as 储蓄率,
          count(*) as 明细条数
        from confirmed_transactions
        where period_month between '2025-01' and '2026-04'
        group by period_month
        order by period_month
        """,
    ),
    (
        "03_支出分类汇总",
        """
        select ct.period_month as 月份, coalesce(c.name, ct.raw_category_snapshot, '未分类') as 分类,
               round(sum(ct.amount), 2) as 金额, count(*) as 条数
        from confirmed_transactions ct
        left join categories c on c.id = ct.category_id
        where ct.transaction_type='expense' and ct.include_in_stats=1 and ct.confirmation_status='confirmed'
          and ct.period_month between '2025-01' and '2026-04'
        group by ct.period_month, 分类
        order by ct.period_month, 金额 desc
        """,
    ),
    (
        "04_收入分类汇总",
        """
        select ct.period_month as 月份, coalesce(c.name, ct.raw_category_snapshot, '未分类') as 分类,
               round(sum(ct.amount), 2) as 金额, count(*) as 条数
        from confirmed_transactions ct
        left join categories c on c.id = ct.category_id
        where ct.transaction_type='income' and ct.include_in_stats=1 and ct.confirmation_status='confirmed'
          and ct.period_month between '2025-01' and '2026-04'
        group by ct.period_month, 分类
        order by ct.period_month, 金额 desc
        """,
    ),
    (
        "05_收支明细抽查",
        """
        select ct.period_month as 月份, ct.transaction_date as 日期, ct.transaction_type as 收支类型,
               round(ct.amount,2) as 金额, coalesce(c.name, ct.raw_category_snapshot, '未分类') as 分类,
               ct.raw_category_snapshot as 原始分类, ct.note as 备注, ct.source_kind as 来源,
               ct.include_in_stats as 是否计入
        from confirmed_transactions ct
        left join categories c on c.id = ct.category_id
        where ct.period_month between '2025-01' and '2026-04'
        order by ct.transaction_date, ct.transaction_type, ct.amount desc
        """,
    ),
    (
        "06_资产月末汇总",
        """
        with asset_months as (
          select period_month, round(sum(amount_cny),2) as asset_gross, count(*) as asset_count
          from monthly_asset_snapshots
          where period_month between '2025-05' and '2026-04' and version_no=1 and status='held'
          group by period_month
        ),
        credit_months as (
          select period_month, round(sum(net_adjustment),2) as credit_net
          from monthly_credit_card_entries
          where period_month between '2025-05' and '2026-04'
          group by period_month
        )
        select
          a.period_month as 月份,
          a.asset_gross as 资产原值明细合计,
          coalesce(c.credit_net, 0) as 信用卡净调整,
          round(a.asset_gross + coalesce(c.credit_net, 0), 2) as Numbers口径资产总额,
          a.asset_count as 资产条数
        from asset_months a
        left join credit_months c on c.period_month = a.period_month
        order by a.period_month
        """,
    ),
    (
        "07_资产月末明细",
        """
        select mas.period_month as 月份, a.name as 资产名称, ac.name as 主资产类别, sub.name as 子类别,
               round(mas.original_amount,2) as 原币金额, mas.currency as 币种,
               round(mas.fx_rate_to_cny,6) as 汇率到人民币, round(mas.amount_cny,2) as 人民币金额,
               mas.note as 备注
        from monthly_asset_snapshots mas
        join assets a on a.id=mas.asset_id
        left join asset_categories ac on ac.id=a.main_asset_category_id
        left join asset_categories sub on sub.id=a.sub_asset_category_id
        where mas.period_month between '2025-05' and '2026-04' and mas.version_no=1
        order by mas.period_month, 人民币金额 desc
        """,
    ),
    (
        "08_投资现金流",
        """
        select ic.period_month as 月份, ic.flow_date as 日期, a.name as 资产名称,
               case ic.flow_type when 'buy' then '买入' when 'sell' then '卖出' when 'dividend' then '分红' else ic.flow_type end as 类型,
               round(ic.amount,2) as 原币金额, ic.currency as 币种, round(ic.amount_cny,2) as 人民币金额,
               ic.source_kind as 来源, ic.note as 备注
        from investment_cashflows ic
        join assets a on a.id=ic.asset_id
        where ic.period_month between '2025-05' and '2026-04'
        order by ic.flow_date, a.name
        """,
    ),
    (
        "09_月度完成记录",
        """
        select period_month as 月份, status as 状态, close_date as 月末日期, confirmed_at as 确认时间, note as 备注
        from monthly_closes order by period_month
        """,
    ),
    (
        "10_异常检查",
        """
        select '未分类已确认收支' as 异常类型, count(*) as 数量, min(period_month) as 最早月份, max(period_month) as 最晚月份
        from confirmed_transactions where category_id is null and confirmation_status='confirmed'
        union all
        select '疑似重复原始账单', count(*), min(substr(transaction_date,1,7)), max(substr(transaction_date,1,7))
        from raw_transactions where potential_duplicate=1 and duplicate_review_status='pending'
        union all
        select '非人民币资产快照', count(*), min(period_month), max(period_month)
        from monthly_asset_snapshots where currency <> 'CNY'
        union all
        select '非人民币投资现金流', count(*), min(period_month), max(period_month)
        from investment_cashflows where currency <> 'CNY'
        """,
    ),
]


def fit_columns(ws):
    for col in range(1, ws.max_column + 1):
        width = min(
            max(10, max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 200) + 1)) + 2),
            46,
        )
        ws.column_dimensions[get_column_letter(col)].width = width


def add_table(ws, rows):
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E9EEF7")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row[key] for key in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fit_columns(ws)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("00_说明")
    for row in [
        ["数据库", str(DB)],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["核对顺序", "02 月度收支汇总 -> 03/04 分类汇总 -> 07 资产月末明细 -> 08 投资现金流 -> 10 异常检查"],
        ["注意", "2025-01 到 2026-04 收支、资产、投资现金流均按 Numbers 历史数据导入；不再参考鲨鱼 CSV 明细。"],
    ]:
        ws.append(row)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 130

    for sheet_name, query in QUERIES:
        ws = wb.create_sheet(sheet_name)
        add_table(ws, conn.execute(query).fetchall())
    conn.close()

    expected = numbers_expected()
    actual = db_actual()
    reconciliation_rows = []
    labels = {
        "income": "收入",
        "expense": "支出",
        "asset_gross": "资产原值",
        "credit_net": "信用卡净调整",
        "asset_total": "Numbers口径资产总额",
    }
    for group in ["income", "expense", "asset_gross", "credit_net", "asset_total"]:
        for period in sorted(set(expected[group]) | set(actual[group])):
            source_value = round(expected[group].get(period, 0.0), 2)
            db_value = round(actual[group].get(period, 0.0), 2)
            diff = round(db_value - source_value, 2)
            reconciliation_rows.append(
                {
                    "项目": labels[group],
                    "月份": period,
                    "Numbers源值": source_value,
                    "数据库值": db_value,
                    "差额": diff,
                    "状态": "一致" if abs(diff) <= TOLERANCE else "不一致",
                }
            )
    ws = wb.create_sheet("11_Numbers对账")
    add_table(ws, reconciliation_rows)
    wb.save(out_path)
    print(out_path)


if __name__ == "__main__":
    main()
